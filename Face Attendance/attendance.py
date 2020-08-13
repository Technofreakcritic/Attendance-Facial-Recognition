import face_recognition
import calendar

import cv2
import datetime
import os
import time
import sys
import pyttsx3
import mysql.connector
import mysql

from plyer    import notification
from openpyxl import Workbook
from openpyxl import load_workbook
from PyQt5.QtWidgets import *

def notification_success():
    notification.notify(
        title='Attendance marked',
        message='Your image has been captured succesfully ',
        timeout=2
        )
def on_button_clicked():
    alert = QMessageBox()
    alert.setText('You clicked the button!')
    alert.exec_()
def closeEvent():
    sys.exit()

# Camera IP Configuration

#******************************************************************************************
# def camera():
    
    # Camera IP Configuration
    # username='admin'
    # password=''
    # ip ='XXX.XXX.XXX.XXX.XXX'
    # port='XXX'

    # ip='http://'+username+password+':@'+ip+':'+port+'/videostream.cgi'

    # # Get a reference to webcam #0 (the default one)
    # video_capture = cv2.VideoCapture(ip)

    # return video_capture
#****************************************************************************************

#MYSQL DB
def database():
    mydb = mysql.connector.connect(
        host="192.168.1.180",
        user="root",
        passwd="",
        database="DBNAME")

#******************************************************************************************
# Storage Location
def storage(path):

    if not os.path.exists(path):
        os.makedirs(path)
        print("New folder created ")
    else:
        print("Folder already exists")  
################################################################################################


# Enable Local Video Camera Capture
video_capture = cv2.VideoCapture(0)


# Time and Date Values

# Load present date and time
now     = datetime.datetime.now()
today   =now.day
month   =now.month

# Date function
get_date = datetime.date.today()
current_date = "{:%b-%d-%Y}".format(get_date)
rent_date    = "{:%b}".format(get_date)

today_date ="{:%d}".format(get_date)

print("Today's date: "+ str(current_date))
print("Today's month: "+ str(rent_date))
print("Today's month: "+ str(month))
print("Today's date: "+ str(today_date))

no_of_days = calendar.monthrange(now.year, now.month)[1]

path = '/home/TFC/Desktop/Face Attendance'

# Voice engine
engine = pyttsx3.init()

#    Check if file exists and create if doesn't exists
exists = os.path.isfile(str(rent_date)+'.xlsx')

storage(path)

global book
global sheet

if (not exists):
    # Create Excel notebook
    print("Doesn't Exist")
    book =  Workbook()
    sheet = book.active
    sheet.cell(row=2, column=2).value = "Employee Name"
    sheet.cell(row=2, column=3).value = "Employee Id"

else:
    print("The file exist \n")
    wb = load_workbook(str(rent_date)+'.xlsx')
    sheet = wb.active


#****************************************************************************


counters = 1

while(no_of_days!=counters):
    sheet.cell(row=2, column=5+counters).value = counters
    counters = counters+1


Em_names = [ "Arshaad", "Sheraz", "Kamil" , "Ykeong", "Adam","Aiman", "Asad"]
Em_ID = [ "10","11","12","13","14","15","16"]


j=len(Em_names)
count=0

while(j!=count):
    sheet.cell(row=count+4, column=2).value = Em_names[count]
    sheet.cell(row=count+4, column=3).value = Em_ID[count]
    count += 1;

location ="Employees"

known_face_encodings = []
known_face_names =  Em_ID


train_dir = os.listdir(location) 

for pics in train_dir:
    #  for human_face in emp_pic:
    face = face_recognition.load_image_file( location  + "/" + pics)
    face_enc = face_recognition.face_encodings(face)[0]
    known_face_encodings.append(face_enc)
    print("Read complete : "+pics)


# Initialize some variables
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True


print("Welcome to HELIOS Inc.")
engine.say('Welcome to HELIOS Inc.')
engine.say('Lets get Started')
engine.runAndWait()

checkmate = False

while True:

    # Grab a single frame of video
    ret, frame = video_capture.read()
    # Resize frame of video to 1/4 size for faster face recognition processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
    rgb_small_frame = small_frame[:, :, ::-1]


    # Only process every other frame of video to save time
    if (process_this_frame):
        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

    face_names = []
    for face_encoding in face_encodings:

        # See if the face is a match for the known face(s)
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        name = "Unknown"

        # If a match was found in known_face_encodings, just use the first one.
        if True in matches:
            first_match_index = matches.index(True)
            name = known_face_names[first_match_index]

             # Assign attendance
            if int(name) in range(1,20):
                sheet.cell(row=int(name)-6, column= 5+get_date.day).value = "Present"
                checkmate= True
            else:
                pass

        face_names.append(name)

    process_this_frame = not process_this_frame

    # Display the results
    for (top, right, bottom, left), name in zip(face_locations, face_names):
           # Scale back up face locations since the frame we detected in was scaled to 1/4 size
           top *= 4
           right *= 4
           bottom *= 4
           left *= 4

           # Draw a box around the face
           cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

           # Draw a label with a name below the face
           cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
           font = cv2.FONT_HERSHEY_DUPLEX
           cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

    # Display the resulting image
    cv2.imshow('HELIOS Attendance ', frame)
    # button_b.clicked.connect(closeEvent)
    # button_b.show()

    # Save Woorksheet as present month
    wb.save(str(rent_date)+'.xlsx')

    # if (checkmate == True):
    #     notification_success()
    #     #time.sleep(5)
    #     #break

    # Hit 'q' on the keyboard to quit!
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release handle to the webcam
video_capture.release()
cv2.destroyAllWindows()
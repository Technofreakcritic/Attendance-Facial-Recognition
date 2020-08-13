import face_recognition
import cv2
import datetime
import os
import time
import sys
import pyttsx3
import mysql.connector
import mysql

from plyer    import notification
from openpyxl import Workbook
from openpyxl import load_workbook
from PyQt5.QtWidgets import *

def notification_success():
    notification.notify(
        title='Attendance marked',
        message='Your image has been captured succesfully ',
        timeout=2
        )
def on_button_clicked():
    alert = QMessageBox()
    alert.setText('You clicked the button!')
    alert.exec_()
def closeEvent():
    sys.exit()

#Camera Connection

#****************************************************************************************
# Camera IP Configuration
username='admin'
password=''
ip ='192.168.1.131'
port='9500'

ip='http://'+username+password+':@'+ip+':'+port+'/videostream.cgi'

# Get a reference to webcam #0 (the default one)
video_capture = cv2.VideoCapture(ip)
#****************************************************************************************

# mydb = mysql.connector.connect(
#   host="192.168.1.180",
#   user="root",
#   passwd="",
#   #database="senzoserver"
# )

#print(mydb)







# Time and Date Values

# Load present date and time
now     = datetime.datetime.now()
today   =now.day
month   =now.month

get_date = datetime.date.today()
current_date = "{:%b-%d-%Y}".format(get_date)
rent_date    = "{:%b}".format(get_date)

print("Today's date: "+ str(current_date)+"\n")

print("Today's month: "+ str(rent_date)+"\n")
print("Today's month: "+ str(month)+"\n")


#******************************************************************************************

# Date function
get_date = datetime.date.today()
#print("{:%b-%d-%Y}".format(get_date))
current_date = "{:%b-%d-%Y}".format(get_date)
################################################
# Storage Location

path = '/home/cofee/Desktop/Attendance'
#path = '/home/cofee/Desktop/Detect_new/'+current_date
if not os.path.exists(path):
    os.makedirs(path)
    print("New folder created ")
else:
    print("Folder already exists")

################################################################################################

# Voice engine
engine = pyttsx3.init()

#    Check if file exists and create if doesn't exists
exists = os.path.isfile('6.xlsx')


global book
global sheet

if (not exists):
    print("Doesn't Exist")
    book =  Workbook()
    sheet = book.active
    sheet.cell(row=2, column=2).value = "Employee Name"
    sheet.cell(row=2, column=3).value = "Employee Id"

else:
    print("The file exist \n")
    wb = load_workbook('6.xlsx')
    sheet = wb.active


#****************************************************************************

x=31
counters = 1
month_entry = datetime.date.today()
#print("Day : " + str(month_entry.day)+"/"+str(month_entry.month)+"/"+str(month_entry.year))

while(x!=counters):
    sheet.cell(row=2, column=5+counters).value = counters
    counters = counters+1


Em_names = [ "Arshaad", "Sheraz",
             "Kamil" , "Ykeong" ,
             "Adam","Aiman",
             "Asad"]

Em_ID = [ "10","11","12","13","14","15","16"]

j=7
count=0

while(j!=count):
    sheet.cell(row=count+4, column=2).value = Em_names[count]
    sheet.cell(row=count+4, column=3).value = Em_ID[count]
    count = count + 1;

# Load images.

image_1 = face_recognition.load_image_file("1.jpg")

image_1_face_encoding = face_recognition.face_encodings(image_1)[0]
image_2 = face_recognition.load_image_file("2.jpg")
image_2_face_encoding = face_recognition.face_encodings(image_2)[0]

image_3 = face_recognition.load_image_file("3.jpg")
image_3_face_encoding = face_recognition.face_encodings(image_3)[0]

image_4 = face_recognition.load_image_file("4.jpg")
image_4_face_encoding = face_recognition.face_encodings(image_4)[0]

image_5 = face_recognition.load_image_file("5.jpg")
image_5_face_encoding = face_recognition.face_encodings(image_5)[0]

image_6 = face_recognition.load_image_file("6.jpg")
image_6_face_encoding = face_recognition.face_encodings(image_6)[0]

image_7 = face_recognition.load_image_file("retard.jpg")
image_7_face_encoding = face_recognition.face_encodings(image_7)[0]

# Create arrays of known face encodings and their names
known_face_encodings = [
        image_1_face_encoding,        image_2_face_encoding,
        image_3_face_encoding,        image_4_face_encoding,
        image_5_face_encoding,        image_6_face_encoding,
        image_7_face_encoding
    ]

known_face_names =  Em_ID
# known_face_names = ["11" , "20", "13" , "14", "15" , "16" ]

# Initialize some variables
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True

app = QApplication([])
button_a = QPushButton('Start')
button_b = QPushButton('Stop')

# Camera IP Configuration

button_a.clicked.connect(on_button_clicked)
button_a.show()
app.exec_()


print("Welcome to Senzo Inc.")
engine.say('Welcome to Senzo Inc.')
engine.say('Click here to Start')
engine.runAndWait()

input("Press Enter to continue...")
checkmate = False


while True:

    # Grab a single frame of video
    ret, frame = video_capture.read()
    # Resize frame of video to 1/4 size for faster face recognition processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
    rgb_small_frame = small_frame[:, :, ::-1]


    # Only process every other frame of video to save time
    if (process_this_frame):
        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

    face_names = []
    for face_encoding in face_encodings:

        # See if the face is a match for the known face(s)
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        name = "Unknown"

        # If a match was found in known_face_encodings, just use the first one.
        if True in matches:
            first_match_index = matches.index(True)
            name = known_face_names[first_match_index]
        #    print(name)
             # Assign attendance
            if int(name) in range(1,20):
                sheet.cell(row=int(name)-6, column=5+month_entry.day).value = "Present"
                checkmate= True
                #sheet.cell(row=int(name), column=2).value = "Present"
            else:
                pass

        face_names.append(name)

    process_this_frame = not process_this_frame

    # Display the results
    for (top, right, bottom, left), name in zip(face_locations, face_names):
           # Scale back up face locations since the frame we detected in was scaled to 1/4 size
           top *= 4
           right *= 4
           bottom *= 4
           left *= 4

           # Draw a box around the face
           cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

           # Draw a label with a name below the face
           cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
           font = cv2.FONT_HERSHEY_DUPLEX
           cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

    # Display the resulting image
    cv2.imshow('Video', frame)
    button_b.clicked.connect(closeEvent)
    button_b.show()

    # Save Woorksheet as present month

    #book.save(str(month)+'.xlsx')
    #wb.save(str(month)+'.xlsx')

    wb.save(str(rent_date)+'.xlsx')

    if (checkmate == True):
        notification_success()
        #time.sleep(5)
        #break

    # Hit 'q' on the keyboard to quit!
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release handle to the webcam
video_capture.release()
cv2.destroyAllWindows()
